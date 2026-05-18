"""Load and chunk documents (PDF, DOCX, XLSX, TXT) for RAG."""

import os
from pathlib import Path
from typing import List

from langchain_community.document_loaders import (
    Docx2txtLoader,
    PyPDFLoader,
    TextLoader,
)
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

from config import CHUNK_OVERLAP, CHUNK_SIZE


def _load_xlsx(file_path: str) -> List[Document]:
    """Read an .xlsx/.xls file into Documents (one per sheet).

    openpyxl reads only cell values; this approach avoids heavier deps like
    `unstructured` or `pandas`. Each sheet becomes a Document with newline-
    separated rows so the text splitter has natural break points.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True, read_only=True)
    docs: List[Document] = []
    for sheet in wb.worksheets:
        rows_text: List[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows_text.append("\t".join(cells))
        if not rows_text:
            continue
        page_content = f"Sheet: {sheet.title}\n" + "\n".join(rows_text)
        docs.append(Document(page_content=page_content, metadata={"sheet": sheet.title}))
    wb.close()
    return docs


class DocumentProcessor:
    """Load documents by type and split into chunks for embedding."""

    def __init__(
        self,
        chunk_size: int = CHUNK_SIZE,
        chunk_overlap: int = CHUNK_OVERLAP,
    ):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separators=["\n\n", "\n", " ", ""],
        )

    def load_documents(self, file_path: str) -> List[Document]:
        """Load a single file into LangChain Documents by extension."""
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".pdf":
            loader = PyPDFLoader(file_path)
        elif suffix in (".docx", ".doc"):
            loader = Docx2txtLoader(file_path)
        elif suffix == ".txt":
            loader = TextLoader(file_path, encoding="utf-8", autodetect_encoding=True)
        elif suffix in (".xlsx", ".xls"):
            return _load_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file type: {suffix}")

        return loader.load()

    def process_documents(self, file_paths: List[str]) -> List[Document]:
        """Load multiple files, add metadata, and split into chunks."""
        all_documents: List[Document] = []

        for file_path in file_paths:
            try:
                docs = self.load_documents(file_path)
                for doc in docs:
                    doc.metadata["source"] = file_path
                    doc.metadata["filename"] = os.path.basename(file_path)
                all_documents.extend(docs)
            except Exception as e:
                raise RuntimeError(f"Error processing {file_path}: {e}") from e

        chunks = self.text_splitter.split_documents(all_documents)
        return chunks
